import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout

def proccess_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    heading=sheet.cell(1,25)
    heading.value = 'Profit Barchart'
    values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=12,max_col=12)

    chart = BarChart(400)
    chart.add_data(values)
    sheet.add_chart(chart,'u2')

    wb.save(f'corrected_{filename}')